from googleapiclient.discovery import build
import json
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import enum
import io
from tqdm import trange
from urllib.parse import urlparse, parse_qs


# version info
VERSION = 3.2

# Todo: add error handling, excel cell size
RETURN_ERR = -1

#input
print("[Info] InfluencerAnalysis V" + str(VERSION))
input_file = open(".\input.txt", "r", encoding="UTF8")
input_data=input_file.readlines()
input_file.close()
excel_dict = {}
for line in input_data:
    key_value = line.strip().split('=')
    if len(key_value)==2:
        excel_dict[key_value[0]] = key_value[1]

DEVELOPER_KEY = excel_dict["DEVELOPER_KEY"]
INPUT_EXCEL = excel_dict["INPUT_EXCEL"]
OUTPUT_EXCEL = excel_dict["OUTPUT_EXCEL"]
INFLUENCER_SHEET = int(excel_dict["INFLUENCER_SHEET"])
VIDEO_SHEET = int(excel_dict["VIDEO_SHEET"])
START_ROW = int(excel_dict["START_ROW"])
START_COL = int(excel_dict["START_COL"])
MAX_RESULT = int(excel_dict["MAX_RESULT"])

# youtube api
YOUTUBE_API_SERVICE_NAME="youtube"
YOUTUBE_API_VERSION="v3"

def make_enum(*sequential, **named):
    enums = dict(zip(sequential, range(len(sequential))), **named)
    return type('Enum', (), enums)

vIndex = make_enum('V_URL', 'V_TITLE', 'VIEW', 'LIKE', 'COMMENTS', 'C_TITLE', 'C_URL', 'CHANNEL_SUBSCRIBER', 'THUMBNAIL')
cIndex = make_enum('URL', 'PROFILE_IMG', 'TITLE', 'SUBSCRIBER', 'POST_VIEW', 'POST_LIKE', 'POST_COMMENT', 'POST_ENGAGE', 'AGE', 'GENDER', 'LOCATION', 'LANGUAGE')


def get_id_from_url(url):
    """Returns Video_ID extracting from the given url of Youtube

    Examples of URLs:
      Valid:
        'http://youtu.be/_lOT2p_FCvA',
        'www.youtube.com/watch?v=_lOT2p_FCvA&feature=feedu',
        'http://www.youtube.com/embed/_lOT2p_FCvA',
        'http://www.youtube.com/v/_lOT2p_FCvA?version=3&amp;hl=en_US',
        'https://www.youtube.com/watch?v=rTHlyTphWP0&index=6&list=PLjeDyYvG6-40qawYNR4juzvSOg-ezZ2a6',
        'youtube.com/watch?v=_lOT2p_FCvA',
        'https://www.youtube.com/channel/UCUbOogiD-4PKDqaJfSOTC0g'
      Invalid:
        'youtu.be/watch?v=_lOT2p_FCvA',
    """
    if url.startswith(('youtu', 'www')):
        url = 'http://' + url

    query = urlparse(url)

    if 'youtube' in query.hostname:
        if (query.path == '/watch') or (query.path == '//watch'):
            return parse_qs(query.query)['v'][0]
        elif query.path.startswith(('/embed/', '/v/', '/channel/')):
            return query.path.split('/')[2]
    elif 'youtu.be' in query.hostname:
        return query.path[1:]
    elif 'instagram' in query.hostname:
        if query.path.startwith('/p/'):
            return query.path.split('/')[2]
        else:
            return query.path.split('/')[1]
    else:
        return RETURN_ERR


def InsertImage(sheet, img_url, row, col):
    image_scale = 10
    response = requests.get(img_url)
    img_file = io.BytesIO(response.content)
    thumbnailImage = Image(img_file)
    thumbnailImage.width /= image_scale
    thumbnailImage.height /= image_scale
    colChar = get_column_letter(col)
    thumbnailImage.anchor = "%s"%colChar + "%s"%row
    sheet.add_image(thumbnailImage)
    #sheet.column_dimensions[colChar].width = thumbnailImage.width
    sheet.row_dimensions[row].height = thumbnailImage.height


def RequestVideoInfo(vID, dev_key):
    VIDEO_SEARCH_URL = "https://www.googleapis.com/youtube/v3/videos?id=" + vID + "&key=" + dev_key + "&part=snippet,statistics&fields=items(id,snippet(channelId, title, thumbnails.high),statistics)"
    response = requests.get(VIDEO_SEARCH_URL).json()
    return response


def RequestChannelInfo(cID, dev_key):
    CHANNEL_SEARCH_URL = "https://www.googleapis.com/youtube/v3/channels?id=" + cID + "&key=" + dev_key + "&part=snippet,statistics&fields=items(id,snippet(title, thumbnails.high),statistics)"
    response = requests.get(CHANNEL_SEARCH_URL).json()
    return response


def RequestChannelContentsInfo(youtube, cID):
    try:
        response = youtube.search().list(
            channelId = cID,
            type = "video",
            order = "date",
            part = "id",
            fields = "items(id)",
            maxResults = MAX_RESULT
        ).execute()
    except Exception as exception:
        print("[Warning] " + str(exception))
        return RETURN_ERR
    return response


def UpdateVideoInfoToExcel(sheet, r, start, data):
    start_c = start - 1
    sheet.cell(row=r, column=start_c + vIndex.V_TITLE).value = '=HYPERLINK("{}", "{}")'.format(data[vIndex.V_URL], data[vIndex.V_TITLE])
    sheet.cell(row=r, column=start_c + vIndex.VIEW).value = round(float(data[vIndex.VIEW]), 2)
    sheet.cell(row=r, column=start_c + vIndex.LIKE).value = round(float(data[vIndex.LIKE]), 2)
    sheet.cell(row=r, column=start_c + vIndex.COMMENTS).value = round(float(data[vIndex.COMMENTS]), 2)
    sheet.cell(row=r, column=start_c + vIndex.C_TITLE).value = '=HYPERLINK("{}", "{}")'.format(data[vIndex.C_URL], data[vIndex.C_TITLE])
    sheet.cell(row=r, column=start_c + vIndex.C_URL).value = data[vIndex.C_URL]
    sheet.cell(row=r, column=start_c + vIndex.CHANNEL_SUBSCRIBER).value = round(float(data[vIndex.CHANNEL_SUBSCRIBER]), 2)
    InsertImage(sheet, data[vIndex.THUMBNAIL], r, start_c + vIndex.THUMBNAIL)


def UpdateChannelInfoToExcel(sheet, r, start, data):
    start_c = start - 1
    InsertImage(sheet, data[cIndex.PROFILE_IMG], r, start_c + cIndex.PROFILE_IMG)
    sheet.cell(row=r, column=start_c + cIndex.TITLE).value = '=HYPERLINK("{}", "{}")'.format(data[cIndex.URL], data[cIndex.TITLE])
    sheet.cell(row=r, column=start_c + cIndex.SUBSCRIBER).value = data[cIndex.SUBSCRIBER]

    sheet.cell(row=r, column=start_c + cIndex.POST_VIEW).value = round(float(data[cIndex.POST_VIEW]), 2)
    sheet.cell(row=r, column=start_c + cIndex.POST_LIKE).value = round(float(data[cIndex.POST_LIKE]), 2)
    sheet.cell(row=r, column=start_c + cIndex.POST_COMMENT).value = round(float(data[cIndex.POST_COMMENT]),2 )
    sheet.cell(row=r, column=start_c + cIndex.POST_ENGAGE).value = round(float(data[cIndex.POST_ENGAGE]), 2)


def GetChannelData(cID, channel_info, channel_contents_info, dev_key):
    arr = json.dumps(channel_info)
    jsonObject = json.loads(arr)
    if ((jsonObject.get('error')) or ('items' not in jsonObject)):
        print("[Warning] response error! : " + cID)
        print(jsonObject['error'])
        return RETURN_ERR
    items = jsonObject['items']
    if len(items) <= 0:
        print("[Error] no items for Channel Data: " + cID)
        return RETURN_ERR
    item = jsonObject['items'][0]
    ret = {}
    ret[cIndex.URL] = "https://www.youtube.com/channel/" + cID
    ret[cIndex.PROFILE_IMG] = ""
    ret[cIndex.TITLE] = ""
    ret[cIndex.SUBSCRIBER] = 0
    ret[cIndex.POST_VIEW] = 0
    ret[cIndex.POST_LIKE] = 0
    ret[cIndex.POST_COMMENT] = 0
    ret[cIndex.POST_ENGAGE] = 0
    
    try:
        ret[cIndex.PROFILE_IMG] = item['snippet']['thumbnails']['high']['url']
        ret[cIndex.TITLE] = item['snippet']['title']
        ret[cIndex.SUBSCRIBER] = item['statistics']['subscriberCount']

        nViewCnt = 0
        nLikeCnt = 0
        nCommentCnt = 0
        nView = 0;
        nLike = 0;
        nComment = 0;
        for content in channel_contents_info.get("items", []):
            if content["id"]["kind"] != "youtube#video":
                print("[Warning] Type is not video!! check the input: " + cID)
                return RETURN_ERR

            vID = content["id"]["videoId"]
            res_json = RequestVideoInfo(vID, dev_key)

            video_info = GetVideoData(vID, res_json, dev_key)
            view = int(video_info[vIndex.VIEW])
            like = int(video_info[vIndex.LIKE])
            comments = int(video_info[vIndex.COMMENTS])

            if view > 0:
                nViewCnt += view
                nView += 1
            if like > 0:
                nLikeCnt += like
                nLike += 1
            if comments > 0:
                nCommentCnt += comments
                nComment += 1
    except Exception as exception:
        print("[Warning]: " + str(exception) + ", Channel ID: " + cID)
        pass
        
    if nView > 0:
        ret[cIndex.POST_VIEW] = nViewCnt / nView
    if nLike > 0:
        ret[cIndex.POST_LIKE] = nLikeCnt / nLike
    if nComment > 0:
        ret[cIndex.POST_COMMENT] = nCommentCnt / nComment
    if nViewCnt > 0:
        ret[cIndex.POST_ENGAGE] = ((nLikeCnt + nCommentCnt) / nViewCnt) * 100
    return ret


def GetVideoData(vID, input_json, dev_key):
    arr = json.dumps(input_json)
    jsonObject = json.loads(arr)
    if ((jsonObject.get('error')) or ('items' not in jsonObject)):
        print("[Warning] response error! : " + vID)
        print(jsonObject['error'])
        return RETURN_ERR
    items = jsonObject['items']
    if len(items) <= 0:
        print("[Warning] no items for Video Data: " + vID)
        return RETURN_ERR
    item = jsonObject['items'][0]
    ret = {}
    ret[vIndex.V_URL] = "https://www.youtube.com/watch?v=" + vID
    ret[vIndex.V_TITLE] = ""
    ret[vIndex.VIEW] = 0
    ret[vIndex.LIKE] = 0
    ret[vIndex.COMMENTS] = 0 
    ret[vIndex.C_TITLE] = ""
    ret[vIndex.C_URL] = ""
    ret[vIndex.CHANNEL_SUBSCRIBER] = 0
    ret[vIndex.THUMBNAIL] = ""    
    
    if ('snippet' in item):
        snippet = item['snippet']
        if ('title' in snippet):
            ret[vIndex.V_TITLE] = snippet['title']
        if ('channelId' in snippet):
            cID = snippet['channelId']
            ret[vIndex.C_URL] = "https://www.youtube.com/channel/" + cID
            channel_info = RequestChannelInfo(cID, dev_key)
            arr = json.dumps(channel_info)
            jsonObject = json.loads(arr)
            if ('items' in jsonObject):
                item_channel = jsonObject['items'][0]
                if ('snippet' in item_channel) and ('title' in item_channel['snippet']):
                    ret[vIndex.C_TITLE] = item_channel['snippet']['title']
                if ('statistics' in item_channel) and ('subscriberCount' in item_channel['statistics']):
                    ret[vIndex.CHANNEL_SUBSCRIBER] = item_channel['statistics']['subscriberCount']
    if ('statistics' in item):
        statistics = item['statistics']
        if ('viewCount' in statistics):
            ret[vIndex.VIEW] = statistics['viewCount']
        if ('likeCount' in statistics):
            ret[vIndex.LIKE] = statistics['likeCount']
        if ('commentCount' in statistics):
            ret[vIndex.COMMENTS] = statistics['commentCount']
    ret[vIndex.THUMBNAIL] = "https://img.youtube.com/vi/" + vID + "/maxresdefault.jpg"

    return ret


def run_VideoAnalysis(sheet, dev_key):
    print("[Info] Running Youtube VideoAnalysis")
    max_row = sheet.max_row + 1
    for row in trange(START_ROW, max_row):
        vURL = sheet.cell(row, START_COL).value
        if vURL == None:
            continue
        vID = get_id_from_url(vURL)
        if vID == RETURN_ERR:
            continue
        res_json = RequestVideoInfo(vID, dev_key)
        df_just_video = GetVideoData(vID, res_json, dev_key)
        if df_just_video == RETURN_ERR:
            continue
        UpdateVideoInfoToExcel(sheet, row, START_COL + 1, df_just_video)


def run_InfluencerAnalysis(sheet, dev_key):
    print("[Info] Running Youtube InfluencerAnalysis")
    youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION, developerKey=dev_key)
    max_row = sheet.max_row + 1
    for row in trange(START_ROW, max_row):
        cURL = sheet.cell(row, START_COL).value
        if cURL == None:
            continue
        cID = get_id_from_url(cURL)
        if cID == RETURN_ERR:
            continue
            # here!
        channel_info = RequestChannelInfo(cID, dev_key)
        channel_contents_info = RequestChannelContentsInfo(youtube, cID)
        if channel_contents_info == RETURN_ERR:
            continue
        df_just_channel = GetChannelData(cID, channel_info, channel_contents_info, dev_key)
        if df_just_channel == RETURN_ERR:
            continue
        UpdateChannelInfoToExcel(sheet, row, START_COL + 1, df_just_channel)



def RequestChannelInfo_Instagram(cID):
    Channel_url = "https://www.instagram.com/" + cID + "/?__a=1"
    response = requests.get(Channel_url).json()


# read excel
xlsx = openpyxl.load_workbook(INPUT_EXCEL)
cSheet = xlsx.worksheets[INFLUENCER_SHEET]
vSheet = xlsx.worksheets[VIDEO_SHEET]
print("[Info] Open input excel: " + INPUT_EXCEL)

# run Analysis
run_VideoAnalysis(vSheet, DEVELOPER_KEY)
run_InfluencerAnalysis(cSheet, DEVELOPER_KEY)

# save excel
xlsx.save(OUTPUT_EXCEL)
print("[Info] Done saving excel: " + OUTPUT_EXCEL)
